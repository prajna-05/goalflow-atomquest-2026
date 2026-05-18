from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
import uuid, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models.models import get_db, User, GoalSheet, Goal, CheckIn, AuditLog, Notification
from utils.scoring import compute_score, compute_overall_score
from ml.predictor import risk_predictor, anomaly_detector, recommend, cache_get, cache_set
from routes.auth import current_user, require_roles

router = APIRouter()

def all_sheets(db: Session):
    return db.query(GoalSheet).options(
        joinedload(GoalSheet.goals).joinedload(Goal.check_ins)
    ).filter(GoalSheet.cycle_year == 2026).all()

def current_quarter():
    m = datetime.now().month
    if 7  <= m <= 9:  return 1, "Q1"
    if 10 <= m <= 12: return 2, "Q2"
    if 1  <= m <= 3:  return 3, "Q3"
    return 4, "Q4"

def days_since(dt):
    if not dt: return 999
    return (datetime.utcnow() - dt).days

# ── Notifications ──────────────────────────────────────────────────────────
@router.get("/notifications")
def notifications(user=Depends(current_user), db: Session = Depends(get_db)):
    return [{"id":n.id,"message":n.message,"type":n.type,"read":n.read,"created_at":n.created_at.isoformat()}
            for n in db.query(Notification).filter(Notification.user_id == user["id"])
                       .order_by(Notification.created_at.desc()).limit(20).all()]

@router.post("/notifications/{nid}/read")
def mark_read(nid: str, user=Depends(current_user), db: Session = Depends(get_db)):
    n = db.query(Notification).filter(Notification.id == nid).first()
    if n: n.read = True; db.commit()
    return {"success": True}

# ── Users ──────────────────────────────────────────────────────────────────
@router.get("/users")
def users(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    return [{"id":u.id,"name":u.name,"email":u.email,"role":u.role,
             "department":u.department,"avatar":u.avatar}
            for u in db.query(User).all()]

# ── Completion dashboard ───────────────────────────────────────────────────
@router.get("/completion")
def completion(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    employees = db.query(User).filter(User.role == "employee").all()
    _, ql = current_quarter()
    rows = []
    for emp in employees:
        sheet = db.query(GoalSheet).options(joinedload(GoalSheet.goals).joinedload(Goal.check_ins))\
                  .filter(GoalSheet.employee_id == emp.id, GoalSheet.cycle_year == 2026).first()
        mgr   = db.query(User).filter(User.id == emp.manager_id).first()
        checkin_done = False
        if sheet:
            checkin_done = all(any(c.quarter == ql for c in g.check_ins) for g in sheet.goals)
        rows.append({
            "employee_id": emp.id, "employee_name": emp.name,
            "department": emp.department, "manager_name": mgr.name if mgr else "N/A",
            "goal_sheet_status": sheet.status if sheet else "not_started",
            "check_in_completed": checkin_done,
            "goals_count": len(sheet.goals) if sheet else 0,
            "overall_score": compute_overall_score(sheet.goals) if sheet else 0,
        })
    return {
        "summary": {
            "total":              len(rows),
            "submitted":          sum(1 for r in rows if r["goal_sheet_status"] in ["pending_approval","approved"]),
            "approved":           sum(1 for r in rows if r["goal_sheet_status"] == "approved"),
            "check_in_completed": sum(1 for r in rows if r["check_in_completed"]),
        },
        "employees": rows
    }

# ── Analytics ──────────────────────────────────────────────────────────────
@router.get("/analytics")
def analytics(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    cached = cache_get("analytics")
    if cached: return cached

    sheets = all_sheets(db)
    dept_scores, thrust_dist, uom_dist = {}, {}, {}

    for s in sheets:
        emp = db.query(User).filter(User.id == s.employee_id).first()
        dept = emp.department if emp else "Unknown"
        dept_scores.setdefault(dept, []).append(compute_overall_score(s.goals))
        for g in s.goals:
            thrust_dist[g.thrust_area] = thrust_dist.get(g.thrust_area, 0) + 1
            uom_dist[g.uom.upper()]    = uom_dist.get(g.uom.upper(), 0) + 1

    dept_data = [{"department":d,"avg_score":round(sum(v)/len(v),1),"count":len(v)} for d,v in dept_scores.items()]

    qoq = []
    for q in ["Q1","Q2","Q3","Q4"]:
        scores = []
        for s in sheets:
            for g in s.goals:
                ci = next((c for c in g.check_ins if c.quarter == q), None)
                if ci: scores.append(compute_score(g.uom, g.uom_direction, g.target, ci.actual))
        qoq.append({"quarter":q,"avg_score":round(sum(scores)/len(scores),1) if scores else 0,"data_points":len(scores)})

    managers = db.query(User).filter(User.role == "manager").all()
    mgr_stats = []
    for mgr in managers:
        team_ids    = [u.id for u in db.query(User).filter(User.manager_id == mgr.id).all()]
        team_sheets = [s for s in sheets if s.employee_id in team_ids]
        approved    = sum(1 for s in team_sheets if s.status == "approved")
        mgr_stats.append({
            "manager_name":  mgr.name,
            "team_size":     len(team_ids),
            "approval_rate": round((approved/len(team_ids))*100,1) if team_ids else 0,
        })

    result = {"dept_data":dept_data,"thrust_dist":thrust_dist,"uom_dist":uom_dist,"qoq":qoq,"mgr_stats":mgr_stats}
    return cache_set("analytics", result)

# ── ML: Risk ──────────────────────────────────────────────────────────────
@router.get("/ml/risk")
def ml_risk(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    cached = cache_get("ml_risk")
    if cached: return cached

    sheets  = [s for s in all_sheets(db) if s.status == "approved"]
    qn, _   = current_quarter()
    results = []
    for sheet in sheets:
        emp        = db.query(User).filter(User.id == sheet.employee_id).first()
        goal_risks = []
        for goal in sheet.goals:
            latest = sorted(goal.check_ins, key=lambda c: c.updated_at)[-1] if goal.check_ins else None
            score  = compute_score(goal.uom, goal.uom_direction, goal.target, latest.actual if latest else "") if latest else 0
            days   = days_since(latest.updated_at) if latest else 90
            risk   = risk_predictor.predict({"uom":goal.uom,"weightage":goal.weightage,"title":goal.title}, score, days, qn)
            goal_risks.append({"goal_title":goal.title,"thrust_area":goal.thrust_area,"weightage":goal.weightage,**risk})
        high = sum(1 for r in goal_risks if r["risk_label"]=="high")
        results.append({
            "employee_id":    sheet.employee_id,
            "employee_name":  emp.name if emp else "",
            "department":     emp.department if emp else "",
            "overall_score":  compute_overall_score(sheet.goals),
            "goal_risks":     goal_risks,
            "high_risk_count":high,
            "employee_risk":  "high" if high>=2 else "medium" if high==1 else "low",
        })
    results.sort(key=lambda x: x["high_risk_count"], reverse=True)
    return cache_set("ml_risk", results)

# ── ML: Anomalies ─────────────────────────────────────────────────────────
@router.get("/ml/anomalies")
def ml_anomalies(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    sheets     = all_sheets(db)
    emp_scores = []
    for sheet in sheets:
        emp = db.query(User).filter(User.id == sheet.employee_id).first()
        qs  = {}
        for goal in sheet.goals:
            for ci in goal.check_ins:
                qs.setdefault(ci.quarter, []).append(compute_score(goal.uom, goal.uom_direction, goal.target, ci.actual))
        avg = lambda q: sum(qs[q])/len(qs[q]) if qs.get(q) else 0
        emp_scores.append({"employee_id":sheet.employee_id,"employee_name":emp.name if emp else "",
                            "q1":avg("Q1"),"q2":avg("Q2"),"q3":avg("Q3"),"q4":avg("Q4")})
    return anomaly_detector.detect(emp_scores)

# ── ML: Recommendations ───────────────────────────────────────────────────
@router.get("/ml/recommendations")
def ml_recommendations(user=Depends(current_user), db: Session = Depends(get_db)):
    emp   = db.query(User).filter(User.id == user["id"]).first()
    sheet = db.query(GoalSheet).options(joinedload(GoalSheet.goals))\
              .filter(GoalSheet.employee_id == user["id"], GoalSheet.cycle_year == 2026).first()
    existing = [{"title":g.title} for g in sheet.goals] if sheet else []
    return recommend(emp.department if emp else "Engineering", existing)

# ── Escalations ───────────────────────────────────────────────────────────
@router.get("/escalations")
def escalations(user=Depends(require_roles("admin")), db: Session = Depends(get_db)):
    results = []
    for emp in db.query(User).filter(User.role == "employee").all():
        sheet = db.query(GoalSheet).filter(GoalSheet.employee_id == emp.id, GoalSheet.cycle_year == 2026).first()
        mgr   = db.query(User).filter(User.id == emp.manager_id).first()
        if not sheet:
            results.append({"type":"no_goals","severity":"high","employee_name":emp.name,
                             "manager_name":mgr.name if mgr else "N/A",
                             "message":f"{emp.name} has not created goals yet","days_overdue":30})
        elif sheet.status == "pending_approval":
            d = days_since(sheet.submitted_at)
            if d > 3:
                results.append({"type":"approval_pending","severity":"high" if d>7 else "medium",
                                 "employee_name":emp.name,"manager_name":mgr.name if mgr else "N/A",
                                 "message":f"{emp.name}'s goals pending approval for {d} days","days_overdue":d})
    return results

# ── Audit log ─────────────────────────────────────────────────────────────
@router.get("/audit")
def audit(user=Depends(require_roles("admin")), db: Session = Depends(get_db)):
    logs = db.query(AuditLog).order_by(AuditLog.performed_at.desc()).limit(200).all()
    return [{"id":l.id,"entity_type":l.entity_type,"entity_id":l.entity_id,
             "action":l.action,"performed_by_name":l.performed_by_user.name if l.performed_by_user else "System",
             "performed_at":l.performed_at.isoformat(),"details":l.details} for l in logs]

# ── Excel export ───────────────────────────────────────────────────────────
@router.get("/export")
def export(user=Depends(require_roles("admin","manager")), db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achievement Report"
    headers  = ["Employee","Department","Thrust Area","Goal","UoM","Target","Weight (%)","Quarter","Actual","Status","Score (%)"]
    hfill    = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill; c.alignment = Alignment(horizontal="center")
    row = 2
    for sheet in all_sheets(db):
        emp = db.query(User).filter(User.id == sheet.employee_id).first()
        for goal in sheet.goals:
            for q in ["Q1","Q2","Q3","Q4"]:
                ci    = next((c for c in goal.check_ins if c.quarter == q), None)
                score = compute_score(goal.uom, goal.uom_direction, goal.target, ci.actual) if ci else "N/A"
                ws.append([emp.name if emp else "","",goal.thrust_area,goal.title,goal.uom,
                            goal.target,goal.weightage,q,ci.actual if ci else "N/A",
                            ci.status if ci else "not_started",score])
                row += 1
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=goalflow_report.xlsx"})
